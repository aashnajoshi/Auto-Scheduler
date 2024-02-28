from openpyxl import Workbook

tt_structure = [["", "1st", "2nd", "3rd", "4th","", "5th", "6th", "7th"], ["Mon"], ["Tues"], ["Wed"], ["Thurs"], ["Fri"]]

def create_timetable(x, tt_structure):
    wb = Workbook()
    ws = wb.active
    for row in tt_structure:
        ws.append(row)

    ws.merge_cells('F1:F6')
    ws.merge_cells('H4:I4')

# to add a condition that if there is a lab lec then 2 cells must be merged using above syntax.

    wb.save(f"{x}_TimeTable.xlsx")
    print(f"{x}_TimeTable.xlsx created!")