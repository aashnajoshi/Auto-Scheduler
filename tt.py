from openpyxl import Workbook

tt_structure = [["", "1st", "2nd", "3rd", "4th","", "5th", "6th", "7th"], ["Mon"], ["Tues"], ["Wed"], ["Thurs"], ["Fri"]]

def create_timetable(x, tt_structure):
    wb = Workbook()
    ws = wb.active
    for row in tt_structure:
        ws.append(row)

    ws.merge_cells('F1:F6')
    ws.merge_cells('H4:I4')
# to add a condition that if there is a lab lec then 2 cells must be merged using above syntax.
    wb.save(f"{x}_TimeTable.xlsx")
    print(f"{x}_TimeTable.xlsx created!")

def get_initial_info():
    # fetch from db: no. names of teachers
    num_of_teachers = int(input("Enter No. of Teachers: "))
    teachers = {}
    for i in range(num_of_teachers):
        name = input(f"Teacher {i+1} Name: ").capitalize()
        
    # subjects taught by specific teacher changes per sem
        num_of_subjects = int(input(f"Number of Subjects taught by {name}: "))
        subjects = [input(f"Subject {j+1} for {name}: ") for j in range(num_of_subjects)]
        teachers[name] = subjects
    return teachers, subjects

teachers_info = get_initial_info()
for name in teachers_info.keys():
    create_timetable(name,tt_structure)